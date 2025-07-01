"""
Gestionnaire de fichiers Excel pour le stockage et l'exportation des données de mesure.
Permet l'organisation, la sauvegarde et la visualisation des données sous forme de graphiques.
"""

import os
import sys
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis

# Ajout du répertoire parent au chemin d'importation pour résoudre les dépendances
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from core.config import config, EXCEL_BASE_DIR

class ExcelHandler:
    """Gère les opérations de fichiers Excel pour le stockage, l'organisation et la visualisation des données de mesure"""
    
    def __init__(self, mode="manual", base_dir=None):
        """
        Initialiser le gestionnaire Excel

        Args:
            mode: Mode de fonctionnement, soit "manual" soit "auto"
            base_dir: Répertoire de base pour les fichiers Excel (par défaut: constants.EXCEL_BASE_DIR)
        """
        self.mode = mode
        self.base_dir = base_dir if base_dir else EXCEL_BASE_DIR
        self.test_folder_path = None
        self.conductance_file = None
        self.co2_temp_humidity_file = None
        self.temp_res_file = None
        
        # Compteurs pour suivre le nombre de séries de données
        self.conductance_series_count = 0
        self.co2_temp_humidity_series_count = 0 
        self.temp_res_series_count = 0

        # Pour suivre les feuilles actives
        self.active_conductance_sheet = None
        self.active_co2_temp_humidity_sheet = None
        self.active_temp_res_sheet = None
        
        # Pour indiquer si un RAZ a été effectué et si une nouvelle feuille est nécessaire
        self.new_conductance_sheet_needed = True
        self.new_co2_temp_humidity_sheet_needed = True  
        self.new_temp_res_sheet_needed = True

        
        # Stocke les données accumulées entre les sessions de réinitialisation (RAZ)
        self.accumulated_conductance_data = {
            'Minutes': [],
            'Temps (s)': [],
            'Conductance (µS)': [],
            'Resistance (Ohms)': []
        }
        
        self.accumulated_co2_temp_humidity_data = {
            'Minutes': [],
            'Temps (s)': [],
            'CO2 (ppm)': [],
            'Température (°C)': [],
            'Humidité (%)': [],
            'deltaC (ppm)': [],
            'masseC (µg)': []
        }
        
        self.accumulated_temp_res_data = {
            'Minutes': [],
            'Temps (s)': [],
            'Température mesurée': [],
            'Tcons': []
        }
    
    def initialize_folder(self):
        """
        Initialise le dossier de test basé sur la date et l'heure actuelles
        
        Returns:
            str: Chemin vers le dossier de test créé
        """
        # Récupérer le répertoire d'exécution de l'application
        if getattr(sys, 'frozen', False):
            # Exécution en tant qu'exécutable compilé
            application_path = os.path.dirname(sys.executable)
        else:
            # Exécution en tant que script Python
            application_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            
        excel_folder_path = os.path.join(application_path, self.base_dir)
        
        # Essayer d'abord le dossier de l'application (où se trouve l'exécutable)
        try:
            if not os.path.exists(excel_folder_path):
                os.makedirs(excel_folder_path, exist_ok=True)
            
            mode_prefix = "Manual" if self.mode == "manual" else "Auto"
            test_folder_name = f"Test-{mode_prefix}-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}"
            self.test_folder_path = os.path.join(excel_folder_path, test_folder_name)
            
            os.makedirs(self.test_folder_path, exist_ok=True)
            print(f"Using application folder for data storage: {self.test_folder_path}")
        except Exception as e:
            # If application folder fails, try user documents folder
            print(f"Failed to create directory in application folder: {e}, trying Documents folder...")
            try:
                # Try Documents folder as fallback
                user_docs = os.path.join(os.path.expanduser('~'), 'Documents', 'ASNR', self.base_dir)
                if not os.path.exists(user_docs):
                    os.makedirs(user_docs, exist_ok=True)
                    
                mode_prefix = "Manual" if self.mode == "manual" else "Auto"
                test_folder_name = f"Test-{mode_prefix}-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}"
                self.test_folder_path = os.path.join(user_docs, test_folder_name)
                
                os.makedirs(self.test_folder_path, exist_ok=True)
                print(f"Using Documents folder for data storage: {self.test_folder_path}")
            except Exception as e:
                # Last resort: use temp directory
                import tempfile
                temp_dir = tempfile.gettempdir()
                print(f"Failed to create directory in Documents folder: {e}, using temp directory: {temp_dir}")

                mode_prefix = "Manual" if self.mode == "manual" else "Auto"
                test_folder_name = f"Test-{mode_prefix}-{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}"
                self.test_folder_path = os.path.join(temp_dir, 'ASNR', self.base_dir, test_folder_name)
                
                os.makedirs(os.path.dirname(self.test_folder_path), exist_ok=True)
                os.makedirs(self.test_folder_path, exist_ok=True)
                print(f"Using temp directory for data storage: {self.test_folder_path}")
        return self.test_folder_path
    
    def initialize_file(self, file_type):
        """
        Initialise un fichier Excel pour un type de données spécifique
        
        Args:
            file_type: Type de fichier de données à initialiser ('conductance', 'co2_temp_humidity', 'temp_res')
            
        Returns:
            str: Chemin vers le fichier initialisé
        """
        if not self.test_folder_path:
            self.initialize_folder()
        
        # Obtenir la date et l'heure actuelles pour le nom du fichier
        current_datetime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        
        # Créer le fichier approprié en fonction du type
        if file_type == "conductance":
            self.conductance_file = os.path.join(self.test_folder_path, f'conductance_{current_datetime}.xlsx')
            return self._create_workbook_with_info(self.conductance_file, file_type)
        
        elif file_type == "co2_temp_humidity":
            self.co2_temp_humidity_file = os.path.join(self.test_folder_path, f'co2_temp_humidity_{current_datetime}.xlsx')
            return self._create_workbook_with_info(self.co2_temp_humidity_file, file_type)
        
        elif file_type == "temp_res":
            self.temp_res_file = os.path.join(self.test_folder_path, f'temperature_resistance_{current_datetime}.xlsx')
            return self._create_workbook_with_info(self.temp_res_file, file_type)
        
        return None
        
    def _create_workbook_with_info(self, file_path, file_type):
        """
        Crée un classeur Excel sans feuille initiale
        
        Args:
            file_path: Chemin du fichier à créer
            file_type: Type de fichier (non utilisé)
            
        Returns:
            str: Chemin du fichier créé
        """
        # Create a workbook with the default sheet
        wb = Workbook()
        
        # Garder la feuille par défaut mais la renommer avec un nom temporaire
        # Cette feuille sera ensuite supprimée lors de l'ajout de vraies données
        ws = wb.active
        ws.title = "_temp"
        
        # Sauvegarder le classeur vide
        wb.save(file_path)
        return file_path
        
    
    def add_sheet_to_excel(self, file_path, sheet_name, data):
        if not os.path.exists(file_path):
            print(f"Erreur: Le fichier {file_path} n'existe pas")
            return False
            
        try:
            wb = load_workbook(file_path)
            
            # Supprimer la feuille temporaire si elle existe
            if "_temp" in wb.sheetnames:
                wb.remove(wb["_temp"])
            
            # Vérifier si la feuille existe déjà
            sheet_exists = sheet_name in wb.sheetnames
            
            # Cas spécial pour "AutoSave": on remplace la feuille existante
            if sheet_exists and sheet_name == "AutoSave":
                # Supprimer la feuille existante pour la recréer
                wb.remove(wb[sheet_name])
                ws = wb.create_sheet(sheet_name)
                print(f"Mise à jour de la feuille 'AutoSave' existante")
            elif sheet_exists:
                # Générer un nom unique en ajoutant un suffixe numérique
                base_name = sheet_name
                counter = 1
                while sheet_name in wb.sheetnames:
                    sheet_name = f"{base_name}_{counter}"
                    counter += 1
                
                # Créer la nouvelle feuille avec le nom généré
                ws = wb.create_sheet(sheet_name)
            else:
                # Créer la nouvelle feuille
                ws = wb.create_sheet(sheet_name)
            
            # Ajouter les données
            for col_num, (key, values) in enumerate(data.items(), 1):
                # Vérifier que les valeurs existent et sont non vides
                if values:
                    ws.cell(row=1, column=col_num, value=key)
                    for row_num, value in enumerate(values, 2):
                        ws.cell(row=row_num, column=col_num, value=value)
            
            
            # Mettre à jour "Essais cumulés" si nécessaire et qu'on a plus d'une feuille de données
            # On ne met pas à jour la feuille cumulée si le nom de feuille est "AutoSave"
            if self._should_create_cumulative_sheet(file_path) and sheet_name != "AutoSave":
                # Sauvegarder avant de mettre à jour la feuille cumulée
                wb.save(file_path)
                self._update_cumulative_sheet(file_path)
                return True
            
            # Sauvegarder les modifications
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"Error adding sheet to Excel file: {e}")
            return False
        
    def _update_cumulative_sheet(self, file_path):
        """
        Met à jour ou crée la feuille 'Essais cumulés' de manière plus robuste
        """
        try:
            # Déterminer le type de données basé sur le nom du fichier
            file_type = None
            if "conductance" in os.path.basename(file_path).lower():
                cumulative_data = self.accumulated_conductance_data
                required_fields = ['Minutes', 'Temps (s)', 'Conductance (µS)', 'Resistance (Ohms)']
                series_count = self.conductance_series_count
                file_type = "conductance"
            elif "co2_temp_humidity" in os.path.basename(file_path).lower():
                cumulative_data = self.accumulated_co2_temp_humidity_data
                required_fields = ['Minutes', 'Temps (s)', 'CO2 (ppm)', 'Température (°C)', 'Humidité (%)']
                series_count = self.co2_temp_humidity_series_count
                file_type = "co2_temp_humidity"
            elif "temperature_resistance" in os.path.basename(file_path).lower():
                cumulative_data = self.accumulated_temp_res_data
                required_fields = ['Minutes', 'Temps (s)', 'Température mesurée', 'Tcons']
                series_count = self.temp_res_series_count
                file_type = "temp_res"
            else:
                return
            
            # Vérifier que nous avons des données accumulées valides
            has_cumulative_data = all(key in cumulative_data for key in required_fields)
            has_data = any(len(cumulative_data.get(key, [])) > 0 for key in required_fields)
            
            if not has_cumulative_data or not has_data:
                return
                
            # Ne créer la feuille "Essais cumulés" que s'il y a plus d'une série de mesures
            if series_count < 2:
                return
                
            # Charger le fichier Excel
            wb = load_workbook(file_path)
            
            # Déterminer les feuilles de données (excluant celle cumulée, la temporaire et AutoSave)
            data_sheets = [s for s in wb.sheetnames if s != "Essais cumulés" and s != "_temp" and s != "AutoSave"]
            
            # Supprimer l'ancienne feuille "Essais cumulés" si elle existe
            if "Essais cumulés" in wb.sheetnames:
                wb.remove(wb["Essais cumulés"])
            
            # Créer la nouvelle feuille "Essais cumulés"
            ws = wb.create_sheet("Essais cumulés")
            
            # Écrire les en-têtes
            for col_num, header in enumerate(required_fields, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # Vérifier si les données accumulées sont cohérentes
            data_lengths = [len(cumulative_data.get(key, [])) for key in required_fields]
            
            if len(set(data_lengths)) > 1:
                # Les données ne sont pas de même longueur - trouver la longueur maximale valide
                valid_lengths = [l for l in data_lengths if l > 0]
                if not valid_lengths:
                    return
                max_rows = max(valid_lengths)
            else:
                # Toutes les données ont la même longueur
                max_rows = data_lengths[0] if data_lengths else 0
            
            # Écrire les données
            for row_num in range(max_rows):
                for col_num, key in enumerate(required_fields, 1):
                    if row_num < len(cumulative_data.get(key, [])):
                        ws.cell(row=row_num+2, column=col_num, value=cumulative_data[key][row_num])
            
            # Sauvegarder les modifications
            wb.save(file_path)
            
        except Exception as e:
            print(f"Erreur lors de la mise à jour de la feuille cumulée: {e}", exc_info=True)

    def raz_conductance_data(self, timeList, conductanceList, resistanceList):
        """Prépare les données pour un nouvel essai sans sauvegarder immédiatement"""
        if timeList and len(timeList) > 0:
            # Ajouter aux données accumulées
            lastTime = 0
            if self.accumulated_conductance_data['Temps (s)']:
                lastTime = self.accumulated_conductance_data['Temps (s)'][-1]
            
            self.accumulated_conductance_data['Minutes'].extend([(lastTime + t) / 60.0 for t in timeList])
            self.accumulated_conductance_data['Temps (s)'].extend([lastTime + t for t in timeList])
            self.accumulated_conductance_data['Conductance (µS)'].extend(conductanceList)
            self.accumulated_conductance_data['Resistance (Ohms)'].extend(resistanceList)
            
            # Marquer qu'une nouvelle feuille sera nécessaire au prochain enregistrement
            self.new_conductance_sheet_needed = True
        
        return True
    
    def raz_co2_temp_humidity_data(self, co2_timestamps, co2_values, temp_timestamps, temp_values, humidity_timestamps, humidity_values):
        """Prépare les données CO2/temp/humidity pour un nouvel essai"""
        if co2_timestamps and len(co2_timestamps) > 0:
            # Ajouter aux données accumulées
            lastTime = 0
            if self.accumulated_co2_temp_humidity_data['Temps (s)']:
                lastTime = self.accumulated_co2_temp_humidity_data['Temps (s)'][-1]
            
            timestamps = co2_timestamps if co2_timestamps else (temp_timestamps if temp_timestamps else humidity_timestamps)
            
            self.accumulated_co2_temp_humidity_data['Minutes'].extend([(lastTime + t) / 60.0 for t in timestamps])
            self.accumulated_co2_temp_humidity_data['Temps (s)'].extend([lastTime + t for t in timestamps])
            self.accumulated_co2_temp_humidity_data['CO2 (ppm)'].extend(co2_values)
            self.accumulated_co2_temp_humidity_data['Température (°C)'].extend(temp_values)
            self.accumulated_co2_temp_humidity_data['Humidité (%)'].extend(humidity_values)
            
            # Marquer qu'une nouvelle feuille sera nécessaire au prochain enregistrement
            self.new_co2_temp_humidity_sheet_needed = True
        
        return True

    def raz_temp_res_data(self, timestamps, temperatures, tcons_values):
        """Prépare les données temp/resistance pour un nouvel essai"""
        if timestamps and len(timestamps) > 0:
            # Ajouter aux données accumulées
            lastTime = 0
            if self.accumulated_temp_res_data['Temps (s)']:
                lastTime = self.accumulated_temp_res_data['Temps (s)'][-1]
            
            self.accumulated_temp_res_data['Minutes'].extend([(lastTime + t) / 60.0 for t in timestamps])
            self.accumulated_temp_res_data['Temps (s)'].extend([lastTime + t for t in timestamps])
            self.accumulated_temp_res_data['Température mesurée'].extend(temperatures)
            self.accumulated_temp_res_data['Tcons'].extend(tcons_values)
            
            # Marquer qu'une nouvelle feuille sera nécessaire au prochain enregistrement
            self.new_temp_res_sheet_needed = True
        
        return True
    
    def save_conductance_data(self, timeList, conductanceList, resistanceList, sheet_name=None):
        """
        Sauvegarde les données de conductance dans le fichier Excel
        
        Args:
            timeList: Liste des timestamps (en secondes)
            conductanceList: Liste des valeurs de conductance (en µS)
            resistanceList: Liste des valeurs de résistance (en Ohms)
            sheet_name: Nom de la feuille à utiliser (ou None pour créer un nom basé sur l'horodatage)
            
        Returns:
            bool: True si la sauvegarde a réussi, False sinon
        """
        if not self.conductance_file:
            self.initialize_file("conductance")
        
        if not timeList or len(timeList) == 0:
            return False
        
        # Déterminer si on doit créer une nouvelle feuille ou utiliser l'existante
        if sheet_name == "AutoSave":
            # Pour les autosaves, toujours utiliser la feuille AutoSave
            pass
        elif not self.new_conductance_sheet_needed and self.active_conductance_sheet:
            # Utiliser la feuille active si pas besoin de nouvelle feuille
            sheet_name = self.active_conductance_sheet
        else:
            # Si aucun nom de feuille n'est fourni ou si une nouvelle est nécessaire, créer un nom
            if sheet_name is None:
                sheet_name = f"Cond_{datetime.now().strftime('%H%M%S')}"
                
            # Mémoriser comme feuille active
            self.active_conductance_sheet = sheet_name
            self.new_conductance_sheet_needed = False
        
        # Définir les données à sauvegarder
        data = {}
        
        # Pour AutoSave, on veut seulement ajouter les nouveaux points non encore enregistrés
        if sheet_name == "AutoSave":
            try:
                wb = load_workbook(self.conductance_file)
                if "AutoSave" in wb.sheetnames:
                    ws = wb["AutoSave"]
                    last_row = ws.max_row
                    if last_row > 1:  # Si la feuille contient déjà des données
                        last_time = ws.cell(row=last_row, column=2).value  # Colonne 2 = Temps (s)
                        if last_time is not None:
                            # Filtrer pour ne prendre que les nouveaux points après le dernier temps enregistré
                            new_indices = [i for i, t in enumerate(timeList) if t > last_time]
                            if new_indices:
                                new_timeList = [timeList[i] for i in new_indices]
                                new_conductanceList = [conductanceList[i] for i in new_indices]
                                new_resistanceList = [resistanceList[i] for i in new_indices]
                                data['Minutes'] = [t / 60.0 for t in new_timeList]
                                data['Temps (s)'] = new_timeList
                                conductanceList = new_conductanceList
                                resistanceList = new_resistanceList
                            else:
                                # Aucun nouveau point à sauvegarder
                                return
                        else:
                            data['Minutes'] = [t / 60.0 for t in timeList]
                            data['Temps (s)'] = timeList
                    else:
                        data['Minutes'] = [t / 60.0 for t in timeList]
                        data['Temps (s)'] = timeList
                else:
                    data['Minutes'] = [t / 60.0 for t in timeList]
                    data['Temps (s)'] = timeList
            except Exception as e:
                print(f"Error checking for existing AutoSave data: {e}")
                data['Minutes'] = [t / 60.0 for t in timeList]
                data['Temps (s)'] = timeList
        else:
            # Pour les autres feuilles, même logique qu'avant
            try:
                if sheet_name in load_workbook(self.conductance_file).sheetnames:
                    wb = load_workbook(self.conductance_file)
                    ws = wb[sheet_name]
                    
                    last_row = ws.max_row
                    if last_row > 1:
                        last_time = ws.cell(row=last_row, column=2).value
                        if last_time is not None:
                            data['Minutes'] = [(last_time + t) / 60.0 for t in timeList]
                            data['Temps (s)'] = [last_time + t for t in timeList]
                        else:
                            data['Minutes'] = [t / 60.0 for t in timeList]
                            data['Temps (s)'] = timeList
                    else:
                        data['Minutes'] = [t / 60.0 for t in timeList]
                        data['Temps (s)'] = timeList
                else:
                    data['Minutes'] = [t / 60.0 for t in timeList]
                    data['Temps (s)'] = timeList
            except Exception as e:
                print(f"Error checking for existing data: {e}")
                data['Minutes'] = [t / 60.0 for t in timeList]
                data['Temps (s)'] = timeList
        
        data['Conductance (µS)'] = conductanceList
        data['Resistance (Ohms)'] = resistanceList
        
        # Pour les données accumulées (continuité entre les tests)
        if sheet_name != "AutoSave":
            # Calculer le dernier point temporel pour la continuité des données cumulées
            lastTime = 0
            if len(self.accumulated_conductance_data['Temps (s)']) > 0:
                lastTime = self.accumulated_conductance_data['Temps (s)'][-1]
            
            # Ajouter les nouvelles données aux données accumulées
            self.accumulated_conductance_data['Minutes'].extend([(lastTime + t) / 60.0 for t in timeList])
            self.accumulated_conductance_data['Temps (s)'].extend([lastTime + t for t in timeList])
            self.accumulated_conductance_data['Conductance (µS)'].extend(conductanceList)
            self.accumulated_conductance_data['Resistance (Ohms)'].extend(resistanceList)
        
        # Sauvegarder ou mettre à jour la feuille
        result = False
        try:
            wb = load_workbook(self.conductance_file)
            
            if sheet_name in wb.sheetnames:
                # La feuille existe, mettre à jour les données
                ws = wb[sheet_name]
                
                # Déterminer la dernière ligne utilisée
                last_row = ws.max_row
                
                # Ajouter les nouvelles données à partir de la dernière ligne
                for i, (minute, second, conductance, resistance) in enumerate(zip(
                    data['Minutes'], data['Temps (s)'], 
                    data['Conductance (µS)'], data['Resistance (Ohms)'])):
                    
                    row = last_row + i + 1  # +1 car nous voulons ajouter après la dernière ligne
                    
                    ws.cell(row=row, column=1, value=minute)
                    ws.cell(row=row, column=2, value=second)
                    ws.cell(row=row, column=3, value=conductance)
                    ws.cell(row=row, column=4, value=resistance)
                
                wb.save(self.conductance_file)
                result = True
            else:
                # La feuille n'existe pas, utiliser la méthode existante pour la créer
                result = self.add_sheet_to_excel(self.conductance_file, sheet_name, data)
                
                # Incrémenter le compteur à chaque nouvelle feuille (sauf AutoSave)
                if sheet_name != "AutoSave":
                    self.conductance_series_count += 1
        except Exception as e:
            print(f"Error saving conductance data: {e}")
            result = self.add_sheet_to_excel(self.conductance_file, sheet_name, data)
        
        return result
    
    def save_co2_temp_humidity_data(self, co2_timestamps, co2_values, temp_timestamps, temp_values, humidity_timestamps, humidity_values, delta_c=None, carbon_mass=None, sheet_name=None):
        """
        Sauvegarde les données de CO2, température et humidité dans le fichier Excel
        
        Args:
            co2_timestamps: Liste des timestamps CO2 (en secondes)
            co2_values: Liste des valeurs CO2 (en ppm)
            temp_timestamps: Liste des timestamps température (en secondes)
            temp_values: Liste des valeurs température (en °C)
            humidity_timestamps: Liste des timestamps humidité (en secondes)
            humidity_values: Liste des valeurs humidité (en %)
            delta_c: Différence de CO2 entre début et fin (en ppm, optionnel)
            carbon_mass: Masse de carbone calculée (en µg, optionnel)
            sheet_name: Nom de la feuille à utiliser (ou None pour créer un nom basé sur l'horodatage)
            
        Returns:
            bool: True si la sauvegarde a réussi, False sinon
        """
        if not self.co2_temp_humidity_file:
            self.initialize_file("co2_temp_humidity")
        
        if not (co2_timestamps or temp_timestamps or humidity_timestamps):
            return False
        
        # Use the first non-empty timestamp list
        timestamps = co2_timestamps if co2_timestamps else (temp_timestamps if temp_timestamps else humidity_timestamps)
        
        # Déterminer si on doit créer une nouvelle feuille ou utiliser l'existante
        if sheet_name == "AutoSave":
            # Pour les autosaves, toujours utiliser la feuille AutoSave
            pass
        elif not self.new_co2_temp_humidity_sheet_needed and self.active_co2_temp_humidity_sheet:
            # Utiliser la feuille active si pas besoin de nouvelle feuille
            sheet_name = self.active_co2_temp_humidity_sheet
        else:
            # Si aucun nom de feuille n'est fourni ou si une nouvelle est nécessaire, créer un nom
            if sheet_name is None:
                sheet_name = f"CO2_{datetime.now().strftime('%H%M%S')}"
                
            # Mémoriser comme feuille active
            self.active_co2_temp_humidity_sheet = sheet_name
            self.new_co2_temp_humidity_sheet_needed = False
        
        # Définir les données à sauvegarder
        data = {}
        
        # Pour AutoSave, on veut seulement ajouter les nouveaux points non encore enregistrés
        if sheet_name == "AutoSave":
            try:
                wb = load_workbook(self.co2_temp_humidity_file)
                if "AutoSave" in wb.sheetnames:
                    ws = wb["AutoSave"]
                    last_row = ws.max_row
                    if last_row > 1:  # Si la feuille contient déjà des données
                        last_time = ws.cell(row=last_row, column=2).value  # Colonne 2 = Temps (s)
                        if last_time is not None:
                            # Filtrer pour ne prendre que les nouveaux points après le dernier temps enregistré
                            new_indices = [i for i, t in enumerate(timestamps) if t > last_time]
                            if new_indices:
                                new_timestamps = [timestamps[i] for i in new_indices]
                                new_co2_values = [co2_values[i] for i in new_indices] if co2_values else []
                                new_temp_values = [temp_values[i] for i in new_indices] if temp_values else []
                                new_humidity_values = [humidity_values[i] for i in new_indices] if humidity_values else []
                                data['Minutes'] = [t / 60.0 for t in new_timestamps]
                                data['Temps (s)'] = new_timestamps
                                timestamps = new_timestamps
                                co2_timestamps = [co2_timestamps[i] for i in new_indices] if co2_timestamps else []
                                temp_timestamps = [temp_timestamps[i] for i in new_indices] if temp_timestamps else []
                                humidity_timestamps = [humidity_timestamps[i] for i in new_indices] if humidity_timestamps else []
                                co2_values = new_co2_values
                                temp_values = new_temp_values
                                humidity_values = new_humidity_values
                            else:
                                # Aucun nouveau point à sauvegarder
                                return
                        else:
                            data['Minutes'] = [t / 60.0 for t in timestamps]
                            data['Temps (s)'] = timestamps
                    else:
                        data['Minutes'] = [t / 60.0 for t in timestamps]
                        data['Temps (s)'] = timestamps
                else:
                    data['Minutes'] = [t / 60.0 for t in timestamps]
                    data['Temps (s)'] = timestamps
            except Exception as e:
                print(f"Error checking for existing AutoSave data: {e}")
                data['Minutes'] = [t / 60.0 for t in timestamps]
                data['Temps (s)'] = timestamps
        else:
            # Pour les autres feuilles, même logique qu'avant
            try:
                if sheet_name in load_workbook(self.co2_temp_humidity_file).sheetnames:
                    wb = load_workbook(self.co2_temp_humidity_file)
                    ws = wb[sheet_name]
                    
                    last_row = ws.max_row
                    if last_row > 1:
                        last_time = ws.cell(row=last_row, column=2).value
                        if last_time is not None:
                            data['Minutes'] = [(last_time + t) / 60.0 for t in timestamps]
                            data['Temps (s)'] = [last_time + t for t in timestamps]
                        else:
                            data['Minutes'] = [t / 60.0 for t in timestamps]
                            data['Temps (s)'] = timestamps
                    else:
                        data['Minutes'] = [t / 60.0 for t in timestamps]
                        data['Temps (s)'] = timestamps
                else:
                    data['Minutes'] = [t / 60.0 for t in timestamps]
                    data['Temps (s)'] = timestamps
            except Exception as e:
                print(f"Error checking for existing data: {e}")
                data['Minutes'] = [t / 60.0 for t in timestamps]
                data['Temps (s)'] = timestamps
        
        data['CO2 (ppm)'] = co2_values
        data['Température (°C)'] = temp_values
        data['Humidité (%)'] = humidity_values
        
        # Pour les données accumulées (continuité entre les tests)
        if sheet_name != "AutoSave":
            # Calculer le dernier point temporel pour la continuité des données cumulées
            lastTime = 0
            if len(self.accumulated_co2_temp_humidity_data['Temps (s)']) > 0:
                lastTime = self.accumulated_co2_temp_humidity_data['Temps (s)'][-1]
            
            # Ajouter les nouvelles données aux données accumulées
            self.accumulated_co2_temp_humidity_data['Minutes'].extend([(lastTime + t) / 60.0 for t in timestamps])
            self.accumulated_co2_temp_humidity_data['Temps (s)'].extend([lastTime + t for t in timestamps])
            self.accumulated_co2_temp_humidity_data['CO2 (ppm)'].extend(co2_values)
            self.accumulated_co2_temp_humidity_data['Température (°C)'].extend(temp_values)
            self.accumulated_co2_temp_humidity_data['Humidité (%)'].extend(humidity_values)
        
        # Sauvegarder ou mettre à jour la feuille
        result = False
        try:
            wb = load_workbook(self.co2_temp_humidity_file)
            
            if sheet_name in wb.sheetnames:
                # La feuille existe, mettre à jour les données
                ws = wb[sheet_name]
                
                # Déterminer la dernière ligne utilisée
                last_row = ws.max_row
                
                # Ajouter les nouvelles données à partir de la dernière ligne
                for i, (minute, second, co2, temp, humidity) in enumerate(zip(
                    data['Minutes'], data['Temps (s)'], 
                    data['CO2 (ppm)'], data['Température (°C)'], data['Humidité (%)'])):
                    
                    row = last_row + i + 1  # +1 car nous voulons ajouter après la dernière ligne
                    
                    ws.cell(row=row, column=1, value=minute)
                    ws.cell(row=row, column=2, value=second)
                    ws.cell(row=row, column=3, value=co2)
                    ws.cell(row=row, column=4, value=temp)
                    ws.cell(row=row, column=5, value=humidity)
                
                # Mettre à jour deltaC et masseC si fournis
                if delta_c is not None:
                    # Trouver la colonne de deltaC ou la créer
                    col_delta = None
                    for col in range(1, ws.max_column + 1):
                        if ws.cell(row=1, column=col).value == "deltaC (ppm)":
                            col_delta = col
                            break
                    
                    if col_delta is None:
                        col_delta = ws.max_column + 1
                        ws.cell(row=1, column=col_delta, value="deltaC (ppm)")
                    
                    ws.cell(row=2, column=col_delta, value=delta_c)
                
                if carbon_mass is not None:
                    # Trouver la colonne de masseC ou la créer
                    col_mass = None
                    for col in range(1, ws.max_column + 1):
                        if ws.cell(row=1, column=col).value == "masseC (µg)":
                            col_mass = col
                            break
                    
                    if col_mass is None:
                        col_mass = ws.max_column + 1
                        ws.cell(row=1, column=col_mass, value="masseC (µg)")
                    
                    ws.cell(row=2, column=col_mass, value=carbon_mass)
                
                wb.save(self.co2_temp_humidity_file)
                result = True
            else:
                # La feuille n'existe pas, utiliser la méthode existante pour la créer
                result = self.add_sheet_to_excel(self.co2_temp_humidity_file, sheet_name, data)
                
                # Incrémenter le compteur à chaque nouvelle feuille (sauf AutoSave)
                if sheet_name != "AutoSave":
                    self.co2_temp_humidity_series_count += 1
        except Exception as e:
            print(f"Error saving CO2/temp/humidity data: {e}")
            result = self.add_sheet_to_excel(self.co2_temp_humidity_file, sheet_name, data)
        
        return result
    
    def save_temp_res_data(self, timestamps, temperatures, tcons_values, sheet_name=None):
        """
        Sauvegarde les données temp/resistance
        
        Args:
            timestamps: Liste des timestamps
            temperatures: Liste des valeurs de température
            tcons_values: Liste des valeurs de consigne de température
            sheet_name: Nom de la feuille à utiliser (ou None pour créer un nom basé sur l'horodatage)
        """
        if not self.temp_res_file:
            self.initialize_file("temp_res")
        
        if not timestamps or not (temperatures or tcons_values):
            return False
        
        # Déterminer si on doit créer une nouvelle feuille ou utiliser l'existante
        if sheet_name == "AutoSave":
            # Pour les autosaves, toujours utiliser la feuille AutoSave
            pass
        elif not self.new_temp_res_sheet_needed and self.active_temp_res_sheet:
            # Utiliser la feuille active si pas besoin de nouvelle feuille
            sheet_name = self.active_temp_res_sheet
        else:
            # Si aucun nom de feuille n'est fourni ou si une nouvelle est nécessaire, créer un nom
            if sheet_name is None:
                sheet_name = f"Temp_{datetime.now().strftime('%H%M%S')}"

            # Mémoriser comme feuille active
            self.active_temp_res_sheet = sheet_name
            self.new_temp_res_sheet_needed = False

        # Définir les données à sauvegarder
        data = {}

        # Pour AutoSave, on veut seulement ajouter les nouveaux points non encore enregistrés
        if sheet_name == "AutoSave":
            try:
                wb = load_workbook(self.temp_res_file)
                if "AutoSave" in wb.sheetnames:
                    ws = wb["AutoSave"]
                    last_row = ws.max_row
                    if last_row > 1:
                        last_time = ws.cell(row=last_row, column=2).value
                        if last_time is not None:
                            # Filtrer pour ne prendre que les nouveaux points après le dernier temps enregistré
                            new_indices = [i for i, t in enumerate(timestamps) if t > last_time]
                            if new_indices:
                                new_timestamps = [timestamps[i] for i in new_indices]
                                new_temperatures = [temperatures[i] for i in new_indices] if temperatures else []
                                new_tcons_values = [tcons_values[i] for i in new_indices] if tcons_values else []
                                data['Minutes'] = [t / 60.0 for t in new_timestamps]
                                data['Temps (s)'] = new_timestamps
                                timestamps = new_timestamps
                                temperatures = new_temperatures
                                tcons_values = new_tcons_values
                            else:
                                # Aucun nouveau point à sauvegarder
                                return
                        else:
                            data['Minutes'] = [t / 60.0 for t in timestamps]
                            data['Temps (s)'] = timestamps
                    else:
                        data['Minutes'] = [t / 60.0 for t in timestamps]
                        data['Temps (s)'] = timestamps
                else:
                    data['Minutes'] = [t / 60.0 for t in timestamps]
                    data['Temps (s)'] = timestamps
            except Exception as e:
                print(f"Error checking for existing AutoSave data: {e}")
                data['Minutes'] = [t / 60.0 for t in timestamps]
                data['Temps (s)'] = timestamps
        else:
            try:
                if sheet_name in load_workbook(self.temp_res_file).sheetnames:
                    wb = load_workbook(self.temp_res_file)
                    ws = wb[sheet_name]

                    last_row = ws.max_row
                    if last_row > 1:
                        last_time = ws.cell(row=last_row, column=2).value
                        if last_time is not None:
                            data['Minutes'] = [(last_time + t) / 60.0 for t in timestamps]
                            data['Temps (s)'] = [last_time + t for t in timestamps]
                        else:
                            data['Minutes'] = [t / 60.0 for t in timestamps]
                            data['Temps (s)'] = timestamps
                    else:
                        data['Minutes'] = [t / 60.0 for t in timestamps]
                        data['Temps (s)'] = timestamps
                else:
                    data['Minutes'] = [t / 60.0 for t in timestamps]
                    data['Temps (s)'] = timestamps
            except Exception as e:
                print(f"Error checking for existing data: {e}")
                data['Minutes'] = [t / 60.0 for t in timestamps]
                data['Temps (s)'] = timestamps

        data['Température mesurée'] = temperatures
        data['Tcons'] = tcons_values
        
        # Pour les données accumulées (continuité entre les tests)
        if sheet_name != "AutoSave":
            # Calculer le dernier point temporel pour la continuité des données cumulées
            lastTime = 0
            if len(self.accumulated_temp_res_data['Temps (s)']) > 0:
                lastTime = self.accumulated_temp_res_data['Temps (s)'][-1]
            
            # Ajouter les nouvelles données aux données accumulées
            self.accumulated_temp_res_data['Minutes'].extend([(lastTime + t) / 60.0 for t in timestamps])
            self.accumulated_temp_res_data['Temps (s)'].extend([lastTime + t for t in timestamps])
            self.accumulated_temp_res_data['Température mesurée'].extend(temperatures)
            self.accumulated_temp_res_data['Tcons'].extend(tcons_values)
        
        # Sauvegarder ou mettre à jour la feuille
        result = False
        try:
            # Vérifier si la feuille existe déjà pour l'ajouter ou la mettre à jour
            wb = load_workbook(self.temp_res_file)
            if sheet_name in wb.sheetnames:
                # La feuille existe, mettre à jour les données
                ws = wb[sheet_name]
                
                # Déterminer la dernière ligne utilisée
                last_row = ws.max_row
                
                # Ajouter les nouvelles données à partir de la dernière ligne
                for i, (minute, second, temp, tcons) in enumerate(zip(
                    data['Minutes'], data['Temps (s)'], 
                    data['Température mesurée'], data['Tcons'])):
                    
                    row = last_row + i + 1  # +1 car nous voulons ajouter après la dernière ligne
                    
                    ws.cell(row=row, column=1, value=minute)
                    ws.cell(row=row, column=2, value=second)
                    ws.cell(row=row, column=3, value=temp)
                    ws.cell(row=row, column=4, value=tcons)
                
                wb.save(self.temp_res_file)
                result = True
            else:
                # La feuille n'existe pas, utiliser la méthode existante pour la créer
                result = self.add_sheet_to_excel(self.temp_res_file, sheet_name, data)
                
                # Incrémenter le compteur à chaque nouvelle feuille (sauf AutoSave)
                if sheet_name != "AutoSave":
                    self.temp_res_series_count += 1
        except Exception as e:
            print(f"Error saving temp/res data: {e}")
            result = self.add_sheet_to_excel(self.temp_res_file, sheet_name, data)
        
        return result
    
    def save_all_data(self, measurement_manager):
        """
        Save all data to Excel files
        
        Args:
            measurement_manager: MeasurementManager instance with data
            
        Returns:
            bool: True if all data was saved successfully, False otherwise
        """
        success = False  # Commence à False, met à True uniquement si des données ont été sauvegardées
        any_data_saved = False
        
        # Save conductance data
        if measurement_manager.timeList and len(measurement_manager.timeList) > 0:
            result = self.save_conductance_data(
                measurement_manager.timeList,
                measurement_manager.conductanceList,
                measurement_manager.resistanceList
            )
            success = result
            any_data_saved = any_data_saved or result
        
        # Save CO2, temperature and humidity data
        has_co2_data = (measurement_manager.timestamps_co2 and len(measurement_manager.timestamps_co2) > 0)
        has_temp_data = (measurement_manager.timestamps_temp and len(measurement_manager.timestamps_temp) > 0)
        has_humidity_data = (measurement_manager.timestamps_humidity and len(measurement_manager.timestamps_humidity) > 0)
        
        if has_co2_data or has_temp_data or has_humidity_data:
            # Get regeneration results (delta_c and carbon_mass) if available
            delta_c = None
            carbon_mass = None
            if hasattr(measurement_manager, 'regeneration_results') and measurement_manager.regeneration_results:
                delta_c = measurement_manager.regeneration_results.get('delta_c')
                carbon_mass = measurement_manager.regeneration_results.get('carbon_mass')
            
            result = self.save_co2_temp_humidity_data(
                measurement_manager.timestamps_co2,
                measurement_manager.values_co2,
                measurement_manager.timestamps_temp,
                measurement_manager.values_temp,
                measurement_manager.timestamps_humidity,
                measurement_manager.values_humidity,
                delta_c,
                carbon_mass
            )
            success = success and result
            any_data_saved = any_data_saved or result
        
        # Save temperature and resistance data
        if measurement_manager.timestamps_res_temp and len(measurement_manager.timestamps_res_temp) > 0:
            result = self.save_temp_res_data(
                measurement_manager.timestamps_res_temp,
                measurement_manager.temperatures,
                measurement_manager.Tcons_values
            )
            success = success and result
            any_data_saved = any_data_saved or result
        
        if any_data_saved:
            print(f"All data saved successfully to {self.test_folder_path}")
        else:
            print("Aucune donnée à sauvegarder")
            success = False
        
        return success
        
    def rename_test_folder(self, new_name):
        """
        Rename the test folder with a custom name
        
        Args:
            new_name: New name for the test folder
            
        Returns:
            bool: True if the folder was renamed successfully, False otherwise
        """
        if not self.test_folder_path or not os.path.exists(self.test_folder_path):
            return False
            
        try:
            # Get parent directory
            parent_dir = os.path.dirname(self.test_folder_path)
            
            # Create new path with the new name
            new_path = os.path.join(parent_dir, new_name)
            
            # Rename the folder
            os.rename(self.test_folder_path, new_path)
            
            # Update the path
            self.test_folder_path = new_path
            
            # Update file paths
            if self.conductance_file:
                filename = os.path.basename(self.conductance_file)
                self.conductance_file = os.path.join(new_path, filename)
                
            if self.co2_temp_humidity_file:
                filename = os.path.basename(self.co2_temp_humidity_file)
                self.co2_temp_humidity_file = os.path.join(new_path, filename)
                
            if self.temp_res_file:
                filename = os.path.basename(self.temp_res_file)
                self.temp_res_file = os.path.join(new_path, filename)
                
            return True
        except Exception as e:
            print(f"Error renaming test folder: {e}")
            return False
        
    def add_charts_to_excel(self, file_path):
        """
        Ajoute des graphiques aux feuilles Excel en fonction du type de données
        
        Args:
            file_path: Chemin du fichier Excel
            
        Returns:
            bool: True si les graphiques ont été ajoutés avec succès, False sinon
        """
        if not os.path.exists(file_path):
            return False
            
        try:
            wb = load_workbook(file_path)
            
            # Déterminer le type de données basé sur le nom du fichier
            if "conductance" in os.path.basename(file_path).lower():
                self._add_conductance_charts(wb)
            elif "co2_temp_humidity" in os.path.basename(file_path).lower():
                self._add_co2_temp_humidity_charts(wb)
            elif "temperature_resistance" in os.path.basename(file_path).lower():
                self._add_temp_res_charts(wb)
                
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"Erreur lors de l'ajout des graphiques Excel: {e}")
            return False

    def _add_conductance_charts(self, workbook):
        """
        Ajoute les graphiques pour les données de conductance
        
        Args:
            workbook: Classeur Excel ouvert
        """
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.axis import ChartLines
        from openpyxl.drawing.line import LineProperties
        from datetime import date
        
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith("Cond_") or sheet_name == "Essais cumulés":
                ws = workbook[sheet_name]
                
                # Supprimer les graphiques existants pour éviter l'accumulation
                charts_to_remove = []
                for chart in ws._charts:
                    if chart.anchor.startswith("H5"):
                        charts_to_remove.append(chart)
                
                for chart in charts_to_remove:
                    ws._charts.remove(chart)
                
                # Créer le graphique principal
                chart = LineChart()
                
                # Titre avec la date actuelle
                today = date.today().strftime("%d/%m/%Y")
                chart.title = f"Essai {today}"
                chart.style = 2
                
                # Configuration de l'axe principal (Conductance)
                chart.y_axis.title = 'Conductance (µS)'
                chart.x_axis.title = 'Temps'
                
                # Forcer l'affichage des axes
                chart.y_axis.delete = False
                chart.x_axis.delete = False
                
                # Activer la grille principale (une seule fois)
                if not hasattr(chart.y_axis, 'majorGridlines') or chart.y_axis.majorGridlines is None:
                    chart.y_axis.majorGridlines = ChartLines()
                if not hasattr(chart.x_axis, 'majorGridlines') or chart.x_axis.majorGridlines is None:
                    chart.x_axis.majorGridlines = ChartLines()
                
                # Configuration des graduations et étiquettes
                chart.y_axis.majorTickMark = "out"
                chart.x_axis.majorTickMark = "out"
                chart.y_axis.minorTickMark = "none"
                chart.x_axis.minorTickMark = "none"
                chart.y_axis.tickLblPos = "nextTo"
                chart.x_axis.tickLblPos = "nextTo"
                
                # Références pour les données (en supposant que vos données commencent à la ligne 2)
                categories = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
                conductance_data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=ws.max_row)
                
                # Ajouter la série Conductance
                chart.add_data(conductance_data, titles_from_data=True)
                chart.set_categories(categories)
                
                # Forcer le format de l'axe X pour qu'il affiche les valeurs numériques
                chart.x_axis.number_format = '0.0'
                
                # Configurer la couleur de la série (bleue comme CO2)
                if len(chart.series) >= 1:
                    # Série Conductance (bleue)
                    line_props_cond = LineProperties(solidFill="4472C4", w=25000)
                    chart.series[0].graphicalProperties.line = line_props_cond
                    chart.series[0].marker = None
                    chart.series[0].smooth = False
                
                # Dimensionner le graphique
                chart.width = 15
                chart.height = 10
                
                # Position du graphique
                ws.add_chart(chart, "H5")
                
                # Debug: vérifier les données
                print(f"Feuille: {sheet_name}")
                print(f"Nombre de lignes: {ws.max_row}")
                print(f"Nombre de colonnes: {ws.max_column}")
                
                # Vérifier les en-têtes de colonnes
                if ws.max_row > 1:
                    print(f"En-têtes: {[ws.cell(1, col).value for col in range(1, min(4, ws.max_column + 1))]}")
                    print(f"Première ligne de données: {[ws.cell(2, col).value for col in range(1, min(4, ws.max_column + 1))]}")

    def _add_co2_temp_humidity_charts(self, workbook):
        """
        Ajoute les graphiques pour les données CO2/température/humidité
        
        Args:
            workbook: Classeur Excel ouvert
        """
        from openpyxl.chart import ScatterChart, Reference, LineChart
        from openpyxl.chart.series import Series
        from openpyxl.chart.axis import ChartLines
        from openpyxl.drawing.line import LineProperties
        from openpyxl.drawing.colors import ColorChoice
        from datetime import date
        
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith("CO2_") or sheet_name == "Essais cumulés":
                ws = workbook[sheet_name]
                
                # Supprimer les graphiques existants pour éviter l'accumulation
                charts_to_remove = []
                for chart in ws._charts:
                    if chart.anchor.startswith("H5"):
                        charts_to_remove.append(chart)
                
                for chart in charts_to_remove:
                    ws._charts.remove(chart)
                
                # Créer le graphique principal
                chart = LineChart()
                
                # Titre avec la date actuelle
                today = date.today().strftime("%d/%m/%Y")
                chart.title = f"Essai {today}"
                chart.style = 21
                
                # Configuration de l'axe principal (CO2)
                chart.y_axis.title = 'CO2 (ppm)'
                chart.x_axis.title = 'Temps'
                
                # Forcer l'affichage des axes
                chart.y_axis.delete = False
                chart.x_axis.delete = False
                
                # Activer la grille principale (une seule fois)
                if not hasattr(chart.y_axis, 'majorGridlines') or chart.y_axis.majorGridlines is None:
                    chart.y_axis.majorGridlines = ChartLines()
                if not hasattr(chart.x_axis, 'majorGridlines') or chart.x_axis.majorGridlines is None:
                    chart.x_axis.majorGridlines = ChartLines()
                
                # Configuration des graduations et étiquettes
                chart.y_axis.majorTickMark = "out"
                chart.x_axis.majorTickMark = "out"
                chart.y_axis.minorTickMark = "none"
                chart.x_axis.minorTickMark = "none"
                chart.y_axis.tickLblPos = "nextTo"
                chart.x_axis.tickLblPos = "nextTo"
                
                # Références pour les données (en supposant que vos données commencent à la ligne 2)
                categories = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
                co2_data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=ws.max_row)
                
                # Ajouter la série CO2
                chart.add_data(co2_data, titles_from_data=True)
                chart.set_categories(categories)
                
                # Forcer le format de l'axe X pour qu'il affiche les valeurs numériques
                chart.x_axis.number_format = '0.0'
                
                # Créer l'axe secondaire pour température et humidité
                temp_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
                hum_data = Reference(ws, min_col=5, min_row=1, max_col=5, max_row=ws.max_row)
                
                chart.add_data(temp_data, titles_from_data=True)
                chart.add_data(hum_data, titles_from_data=True)
                
                # Configuration de l'axe secondaire (droite)
                if len(chart.series) >= 2:
                    # Assigner les séries température et humidité à l'axe secondaire
                    chart.series[1].yAxisId = 200
                    chart.series[2].yAxisId = 200
                
                # Créer l'axe Y secondaire
                from openpyxl.chart.axis import NumericAxis
                chart.y_axis.axId = 100
                chart.y_axis.crosses = "autoZero"
                
                # Axe secondaire avec configuration complète
                ax2 = NumericAxis(axId=200)
                ax2.title = "Température (°C) et Humidité (%)"
                ax2.crosses = "max"
                ax2.majorTickMark = "out"
                ax2.minorTickMark = "none"
                ax2.tickLblPos = "nextTo"
                ax2.delete = False
                
                chart.y_axis_2 = ax2
                
                # Configurer les couleurs des séries
                if len(chart.series) >= 3:
                    # Série CO2 (bleue)
                    line_props_co2 = LineProperties(solidFill="4472C4", w=25000)
                    chart.series[0].graphicalProperties.line = line_props_co2
                    chart.series[0].marker = None
                    chart.series[0].smooth = False
                    
                    # Série Température (rouge)
                    line_props_temp = LineProperties(solidFill="E15759", w=25000)
                    chart.series[1].graphicalProperties.line = line_props_temp
                    chart.series[1].marker = None
                    chart.series[1].smooth = False
                    
                    # Série Humidité (verte)
                    line_props_hum = LineProperties(solidFill="70AD47", w=25000)
                    chart.series[2].graphicalProperties.line = line_props_hum
                    chart.series[2].marker = None
                    chart.series[2].smooth = False
                
                # Dimensionner le graphique
                chart.width = 15
                chart.height = 10
                
                # Position du graphique
                ws.add_chart(chart, "H5")
                
                # Debug: vérifier les données
                print(f"Feuille: {sheet_name}")
                print(f"Nombre de lignes: {ws.max_row}")
                print(f"Nombre de colonnes: {ws.max_column}")
                
                # Vérifier les en-têtes de colonnes
                if ws.max_row > 1:
                    print(f"En-têtes: {[ws.cell(1, col).value for col in range(1, min(6, ws.max_column + 1))]}")
                    print(f"Première ligne de données: {[ws.cell(2, col).value for col in range(1, min(6, ws.max_column + 1))]}")
                
    def _add_temp_res_charts(self, workbook):
        """
        Ajoute les graphiques pour les données température/résistance
        
        Args:
            workbook: Classeur Excel ouvert
        """
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.axis import ChartLines
        from openpyxl.drawing.line import LineProperties
        from datetime import date
        
        for sheet_name in workbook.sheetnames:
            if sheet_name.startswith("Temp_") or sheet_name == "Essais cumulés":
                ws = workbook[sheet_name]
                
                # Supprimer les graphiques existants pour éviter l'accumulation
                charts_to_remove = []
                for chart in ws._charts:
                    if chart.anchor.startswith("H5"):
                        charts_to_remove.append(chart)
                
                for chart in charts_to_remove:
                    ws._charts.remove(chart)
                
                # Créer le graphique principal
                chart = LineChart()
                
                # Titre avec la date actuelle
                today = date.today().strftime("%d/%m/%Y")
                chart.title = f"Essai {today}"
                chart.style = 2
                
                # Configuration de l'axe principal (Température)
                chart.y_axis.title = 'Température (°C)'
                chart.x_axis.title = 'Temps'
                
                # Forcer l'affichage des axes
                chart.y_axis.delete = False
                chart.x_axis.delete = False
                
                # Activer la grille principale (une seule fois)
                if not hasattr(chart.y_axis, 'majorGridlines') or chart.y_axis.majorGridlines is None:
                    chart.y_axis.majorGridlines = ChartLines()
                if not hasattr(chart.x_axis, 'majorGridlines') or chart.x_axis.majorGridlines is None:
                    chart.x_axis.majorGridlines = ChartLines()
                
                # Configuration des graduations et étiquettes
                chart.y_axis.majorTickMark = "out"
                chart.x_axis.majorTickMark = "out"
                chart.y_axis.minorTickMark = "none"
                chart.x_axis.minorTickMark = "none"
                chart.y_axis.tickLblPos = "nextTo"
                chart.x_axis.tickLblPos = "nextTo"
                
                # Références pour les données (en supposant que vos données commencent à la ligne 2)
                categories = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
                temp_data = Reference(ws, min_col=3, min_row=1, max_col=3, max_row=ws.max_row)
                tcons_data = Reference(ws, min_col=4, min_row=1, max_col=4, max_row=ws.max_row)
                
                # Ajouter les séries
                chart.add_data(temp_data, titles_from_data=True)
                chart.add_data(tcons_data, titles_from_data=True)
                chart.set_categories(categories)
                
                # Forcer le format de l'axe X pour qu'il affiche les valeurs numériques
                chart.x_axis.number_format = '0.0'
                
                # Configurer les couleurs des séries
                if len(chart.series) >= 2:
                    # Série Température (rouge)
                    line_props_temp = LineProperties(solidFill="E15759", w=25000)
                    chart.series[0].graphicalProperties.line = line_props_temp
                    chart.series[0].marker = None
                    chart.series[0].smooth = False
                    
                    # Série Tcons (bleue)
                    line_props_tcons = LineProperties(solidFill="4472C4", w=25000)
                    chart.series[1].graphicalProperties.line = line_props_tcons
                    chart.series[1].marker = None
                    chart.series[1].smooth = False
                
                # Dimensionner le graphique
                chart.width = 15
                chart.height = 10
                
                # Position du graphique
                ws.add_chart(chart, "H5")
                
                # Debug: vérifier les données
                print(f"Feuille: {sheet_name}")
                print(f"Nombre de lignes: {ws.max_row}")
                print(f"Nombre de colonnes: {ws.max_column}")
                
                # Vérifier les en-têtes de colonnes
                if ws.max_row > 1:
                    print(f"En-têtes: {[ws.cell(1, col).value for col in range(1, min(5, ws.max_column + 1))]}")
                    print(f"Première ligne de données: {[ws.cell(2, col).value for col in range(1, min(5, ws.max_column + 1))]}")
                    
                    
    def _should_create_cumulative_sheet(self, file_path):
        """Détermine si une feuille 'Essais cumulés' doit être créée"""
        if "conductance" in os.path.basename(file_path).lower():
            return self.conductance_series_count >= 2
        elif "co2_temp_humidity" in os.path.basename(file_path).lower():
            return self.co2_temp_humidity_series_count >= 2
        elif "temperature_resistance" in os.path.basename(file_path).lower():
            return self.temp_res_series_count >= 2
        return False
        